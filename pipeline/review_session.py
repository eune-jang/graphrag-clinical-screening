"""
실제 reviewer 시뮬레이션 — 검수 5개 항목을 기준으로 한 trial을 walk-through하며
판단을 xlsx로 산출.

REVIEW.md의 Step 2-5를 그대로 시행:
  Step 2 (Q0.1)        : 전체 trial 트리아지 — 출력 시트 "00_Triage"
  Step 3 (Q0.4)        : trial 단위 criterion 목록 + 판단 — 시트 "01_Criteria"
  Step 4               : 각 criterion의 5개 항목별 평가 (① ~ ⑤)
  Step 5               : 통계 요약 + 자동화 후보 — 시트 "02_Summary"

사용법:
  python pipeline/review_session.py --trial NCT03425643
  python pipeline/review_session.py --trial NCT03425643 --output results/review_NCT03425643.xlsx
"""
from __future__ import annotations
import argparse
import json
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dotenv import load_dotenv

_HERE = Path(__file__).resolve().parent
load_dotenv(_HERE.parent / ".env")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from neo4j import GraphDatabase


# ── Styles ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SEV_HIGH = PatternFill(start_color="F4B7B7", end_color="F4B7B7", fill_type="solid")
SEV_MED = PatternFill(start_color="FCE5B6", end_color="FCE5B6", fill_type="solid")
SEV_LOW = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
SEV_OK = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, n_cols: int):
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _autosize(ws, max_width: int = 60):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value:
                length = max(len(line) for line in str(cell.value).split("\n"))
                max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


# ── Per-criterion analysis (5 review items) ──────────────────────────

# subtype × relation_type 가 어울리는 쌍 — 03_validate_annotation의 RELATION_SUBTYPE_MAP과 동기화
RELATION_SUBTYPE_MAP = {
    "REQUIRES_BIOMARKER":     {"Biomarker"},
    "REQUIRES_CONDITION":     {"Condition", "Stage"},
    "EXCLUDES_CONDITION":     {"Condition", "Stage"},
    "REQUIRES_TREATMENT":     {"Drug"},
    "EXCLUDES_TREATMENT":     {"Drug"},
    "EXCLUDES_COMEDICATION":  {"Drug"},
    "REQUIRES_PROCEDURE":     {"Procedure"},
    "EXCLUDES_PROCEDURE":     {"Procedure"},
    "REQUIRES_STATUS":        {"Observation", "Condition"},
    "EXCLUDES_STATUS":        {"Observation", "Condition"},
    "INCLUDES_EXCEPTION":     {"Condition", "Drug", "Procedure", "Observation", "Stage"},
    "HAS_VALUE":              {"Observation", "Condition"},
    "HAS_TEMPORAL":           {"Drug", "Condition", "Procedure", "Observation"},
}


def assess_criterion(crit: dict, children: list, all_criteria_by_id: dict) -> dict:
    """5개 항목에 대해 reviewer 판단을 생성.

    각 항목에 대해:
      severity: 'high' (명확한 결함) / 'med' (검토 필요) / 'low' (작은 보강 여지) / 'ok'
      finding:  관찰한 내용 (문장)
      action:   권장 조치 (간결)
    """
    cid = crit["criterion_id"]
    text = crit.get("text", "") or ""
    sc = crit.get("semantic_category")
    pr = crit.get("parent_role")
    cl = crit.get("child_logic")
    pid = crit.get("parent_criterion_id")
    relations = crit.get("relations", []) or []
    valid = crit.get("_validation", {}) or {}
    issues = valid.get("issues", [])

    out: dict[str, dict] = {}

    # ───── ① Criterion 분해 구조 ──────────────────────────────────
    notes_1 = []
    sev_1 = "ok"
    if pr == "composite_split":
        if not children:
            notes_1.append(f"composite_split인데 자녀 없음 (C1)")
            sev_1 = "high"
        elif len(children) < 2:
            notes_1.append(f"composite_split인데 자녀 {len(children)}개 (<2)")
            sev_1 = "high"
        else:
            notes_1.append(f"composite_split + {len(children)}개 자녀 (logic={cl or 'AND default'})")
    elif pr == "macro_aggregate":
        if not children:
            notes_1.append(f"macro_aggregate인데 자녀 없음 (C1)")
            sev_1 = "high"
        else:
            notes_1.append(f"macro_aggregate + {len(children)}개 자녀 (organ function 묶음)")
    elif pr == "nested_exception_parent":
        # carve-out이 self나 children에 있나
        self_carve = any(r.get("relation_type") == "INCLUDES_EXCEPTION" for r in relations)
        child_carve = any(
            any(r.get("relation_type") == "INCLUDES_EXCEPTION" for r in (ch.get("relations") or []))
            for ch in children
        )
        if not (self_carve or child_carve):
            notes_1.append("nested_exception_parent인데 INCLUDES_EXCEPTION 없음 (C2)")
            sev_1 = "high"
        else:
            loc = "self" if self_carve else "children"
            notes_1.append(f"nested_exception_parent + carve-out on {loc}")
    elif pid:
        parent = all_criteria_by_id.get(pid)
        if not parent:
            notes_1.append(f"parent_criterion_id 가리키는 부모 존재 안 함: {pid}")
            sev_1 = "high"
        else:
            # 자녀 text가 부모 text 안에 포함되나
            if text and text not in parent.get("text", ""):
                # fuzzy: token recall
                ptok = set(parent.get("text", "").lower().split())
                ctok = set(text.lower().split())
                recall = len(ctok & ptok) / max(len(ctok), 1)
                if recall < 0.7:
                    notes_1.append(f"자녀 text가 부모와 token recall {recall:.0%} (paraphrase 의심)")
                    sev_1 = "med"
                else:
                    notes_1.append(f"자녀 (부모: {pid}, 의미 일치)")
            else:
                notes_1.append(f"자녀 (부모: {pid}, 직접 substring)")
    else:
        notes_1.append("단일 leaf criterion (분해 없음)")

    out["①"] = {"severity": sev_1, "finding": " | ".join(notes_1) if notes_1 else "(no finding)"}

    # ───── ② Criterion 메타 분류 ──────────────────────────────────
    notes_2 = []
    sev_2 = "ok"
    if not sc:
        notes_2.append("semantic_category null")
        sev_2 = "high"
    else:
        # 키워드 휴리스틱
        tl = text.lower()
        heuristic_hint = None
        if sc == "demographic":
            demo_keys = ("age", "year", "old", "gender", "male", "female",
                         "contracepti", "pregnan", "childbearing", "wocbp",
                         "sperm", "fertili", "abstinen", "breastfeed")
            if not any(k in tl for k in demo_keys):
                heuristic_hint = "demographic 분류인데 텍스트에 demographic 단서 없음"
        elif sc == "comorbidity":
            if "history" not in tl and "active" not in tl and "active" not in tl:
                heuristic_hint = None  # 너무 다양해서 휴리스틱 어려움
        elif sc == "performance_status":
            if "ecog" not in tl and "performance" not in tl and "karnofsky" not in tl:
                heuristic_hint = "performance_status인데 ECOG/Karnofsky 언급 없음 — 재분류 검토"
        elif sc == "biomarker":
            if "egfr" not in tl and "alk" not in tl and "kras" not in tl and "ros1" not in tl and "pd-l1" not in tl and "mutation" not in tl and "biomarker" not in tl:
                heuristic_hint = "biomarker 분류인데 명시적 marker 언급 없음"
        if heuristic_hint:
            notes_2.append(heuristic_hint)
            sev_2 = "med"
        else:
            notes_2.append(f"semantic_category={sc} (text와 부합)")
    cs = crit.get("cohort_scope")
    if cs:
        notes_2.append(f"cohort_scope={cs}")

    out["②"] = {"severity": sev_2, "finding": " | ".join(notes_2)}

    # ───── ③ Cross-layer relation 식별 ────────────────────────────
    notes_3 = []
    sev_3 = "ok"
    if not relations and not pr:
        notes_3.append("leaf criterion인데 relation 0개 (추출 누락 의심)")
        sev_3 = "med"
    else:
        rel_summary = Counter()
        for r in relations:
            rt = r.get("relation_type")
            rel_summary[rt] += 1
            subtype = r.get("target_subtype")
            allowed = RELATION_SUBTYPE_MAP.get(rt)
            if allowed is not None and subtype and subtype not in allowed:
                notes_3.append(f"{rt}→{subtype} 부정합 (R1)")
                sev_3 = "high"
        if "span_not_in_text" in [i.split(":")[0] for i in issues]:
            notes_3.append("span_not_in_text (R2)")
            sev_3 = "high" if sev_3 != "high" else sev_3
        if not notes_3:
            types_str = ", ".join(f"{rt}({n})" for rt, n in rel_summary.most_common(5))
            notes_3.append(f"{len(relations)}개 relation: {types_str}")

    out["③"] = {"severity": sev_3, "finding": " | ".join(notes_3) if notes_3 else "(no relations)"}

    # ───── ④ Relation 속성 완전성 ─────────────────────────────────
    notes_4 = []
    sev_4 = "ok"
    for r in relations:
        rt = r.get("relation_type")
        props = r.get("properties", {}) or {}
        # 필수 키 누락 — validator와 동일 기준 (None/"" 만 missing)
        if rt == "HAS_VALUE":
            missing = [k for k in ("operator", "value") if props.get(k) in (None, "")]
            if missing:
                notes_4.append(f"HAS_VALUE missing {missing} (R4)")
                sev_4 = "high"
            # value=0 suspicious — range 의 하한만 추출했거나 indefinite 의 placeholder 의심
            if props.get("value") == 0 and "0" not in text[:200]:
                notes_4.append(f"HAS_VALUE value=0인데 텍스트에 '0' 명시 없음 — 범위 상한 누락 또는 placeholder 의심")
                sev_4 = "med" if sev_4 != "high" else sev_4
        elif rt == "HAS_TEMPORAL":
            missing = [k for k in ("operator", "value", "unit", "anchor") if props.get(k) in (None, "")]
            if missing:
                notes_4.append(f"HAS_TEMPORAL missing {missing} (R3)")
                sev_4 = "high"
            if props.get("value") == 0:
                notes_4.append(f"HAS_TEMPORAL value=0 — indefinite period 의심, 텍스트 의미 확인 필요")
                sev_4 = "med" if sev_4 != "high" else sev_4
        elif rt == "REQUIRES_BIOMARKER" and not r.get("biomarker_details"):
            notes_4.append("REQUIRES_BIOMARKER missing biomarker_details")
            sev_4 = "high"
        elif rt == "INCLUDES_EXCEPTION":
            if not props.get("exception_type"):
                notes_4.append("INCLUDES_EXCEPTION missing exception_type")
                sev_4 = "med"
        if props.get("alternative_constraint"):
            notes_4.append(f"alt_constraint 있음 (의미 검토)")
    if not notes_4:
        notes_4.append("필수 속성 모두 채워짐")

    out["④"] = {"severity": sev_4, "finding": " | ".join(notes_4)}

    # ───── ⑤ Concept 정규화 (preferred_name 일관성) ───────────────
    # 이 단계는 단일 trial 내에선 약함 — 검수자가 cross-trial 비교 필요 표시만
    notes_5 = []
    sev_5 = "ok"
    pnames = [(r.get("target_preferred_name") or "").strip() for r in relations]
    pnames = [p for p in pnames if p]
    if pnames:
        # 짧은 placeholder 의심
        too_short = [p for p in pnames if len(p) < 4]
        if too_short:
            notes_5.append(f"짧은 preferred_name: {too_short}")
            sev_5 = "med"
        # 자유 텍스트 같은 긴 placeholder 의심
        too_long = [p for p in pnames if len(p) > 80]
        if too_long:
            notes_5.append(f"긴 placeholder 의심: {[p[:60]+'...' for p in too_long]}")
            sev_5 = "med"
        if not notes_5:
            notes_5.append(f"{len(set(pnames))}개 unique preferred_name (cross-trial dedup은 ⑤ Cypher 5.1로)")
    else:
        notes_5.append("preferred_name 없음")

    out["⑤"] = {"severity": sev_5, "finding": " | ".join(notes_5)}

    return out


# ── Main: build xlsx ──────────────────────────────────────────────────

def _sev_fill(sev: str) -> PatternFill | None:
    return {"high": SEV_HIGH, "med": SEV_MED, "low": SEV_LOW, "ok": SEV_OK}.get(sev)


def build_xlsx(trial_data: dict, neo4j_summary: dict, out_path: Path) -> None:
    wb = Workbook()

    # ─── Sheet 1: Triage (30 trial 단위) ──────────────────────────
    ws0 = wb.active
    ws0.title = "00_Triage"
    ws0.append(["#", "Trial NCT", "Acronym", "Criteria", "crit_fail",
                "Relations", "rel_fail", "fail_rate", "Note"])
    _style_header(ws0, 9)
    for i, row in enumerate(neo4j_summary["per_trial"], start=2):
        ws0.append([
            i - 1,
            row["nct"],
            row.get("acronym") or "",
            row["n_crit"],
            row["crit_fail"],
            row["n_rel"],
            row["rel_fail"],
            f"{row['rel_fail']/max(row['n_rel'],1)*100:.1f}%",
            "검수 대상" if row["nct"] == trial_data["trial_id"] else "",
        ])
        if row["nct"] == trial_data["trial_id"]:
            for col in range(1, 10):
                ws0.cell(row=i, column=col).fill = PatternFill(start_color="FFF2CC",
                                                               end_color="FFF2CC", fill_type="solid")
    _autosize(ws0, max_width=30)

    # ─── Sheet 2: Criteria Review (criterion 단위) ──────────────────
    ws1 = wb.create_sheet("01_Criteria")
    ws1.append([
        "criterion_id", "type", "parent_role", "child_logic", "parent_id",
        "n_rel", "text",
        "① 분해", "① finding",
        "② 메타", "② finding",
        "③ relation", "③ finding",
        "④ 속성", "④ finding",
        "⑤ 정규화", "⑤ finding",
        "Reviewer 코멘트 (수기 입력)",
        "Suggested action",
    ])
    _style_header(ws1, 19)
    ws1.freeze_panes = "C2"

    criteria = trial_data["criteria"]
    by_id = {c["criterion_id"]: c for c in criteria}
    children_of: dict[str, list] = {}
    for c in criteria:
        pid = c.get("parent_criterion_id")
        if pid:
            children_of.setdefault(pid, []).append(c)

    for c in sorted(criteria, key=lambda x: x["criterion_id"]):
        cid = c["criterion_id"]
        children = children_of.get(cid, [])
        assess = assess_criterion(c, children, by_id)
        n_rel = len(c.get("relations") or [])
        ws1.append([
            cid,
            c.get("type"),
            c.get("parent_role") or "",
            c.get("child_logic") or "",
            c.get("parent_criterion_id") or "",
            n_rel,
            (c.get("text") or "")[:140],
            assess["①"]["severity"], assess["①"]["finding"],
            assess["②"]["severity"], assess["②"]["finding"],
            assess["③"]["severity"], assess["③"]["finding"],
            assess["④"]["severity"], assess["④"]["finding"],
            assess["⑤"]["severity"], assess["⑤"]["finding"],
            "", "",
        ])
        row_idx = ws1.max_row
        # color the severity cells
        for cat_idx, cat in enumerate(["①", "②", "③", "④", "⑤"]):
            col_sev = 8 + cat_idx * 2  # 8, 10, 12, 14, 16
            cell = ws1.cell(row=row_idx, column=col_sev)
            fill = _sev_fill(assess[cat]["severity"])
            if fill:
                cell.fill = fill
                ws1.cell(row=row_idx, column=col_sev + 1).fill = fill

    # column widths
    widths = {1: 18, 2: 10, 3: 22, 4: 10, 5: 18, 6: 6, 7: 50,
              8: 7, 9: 50, 10: 7, 11: 35, 12: 7, 13: 45,
              14: 7, 15: 35, 16: 7, 17: 30, 18: 30, 19: 30}
    for col_idx, w in widths.items():
        ws1.column_dimensions[get_column_letter(col_idx)].width = w
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    # ─── Sheet 3: Summary ─────────────────────────────────────────
    ws2 = wb.create_sheet("02_Summary")

    # severity 분포
    sev_counter = {cat: Counter() for cat in ["①", "②", "③", "④", "⑤"]}
    for c in criteria:
        a = assess_criterion(c, children_of.get(c["criterion_id"], []), by_id)
        for cat in sev_counter:
            sev_counter[cat][a[cat]["severity"]] += 1

    ws2.append(["검수 항목", "high", "med", "low", "ok", "합계"])
    _style_header(ws2, 6)
    cat_labels = {
        "①": "① Criterion 분해 구조",
        "②": "② Criterion 메타 분류",
        "③": "③ Cross-layer relation 식별",
        "④": "④ Relation 속성 완전성",
        "⑤": "⑤ Concept 정규화",
    }
    for cat, label in cat_labels.items():
        c = sev_counter[cat]
        ws2.append([label, c["high"], c["med"], c["low"], c["ok"], sum(c.values())])

    ws2.append([])
    ws2.append(["Trial 정보"])
    ws2.append(["NCT ID", trial_data["trial_id"]])
    ws2.append(["Acronym", trial_data.get("trial_acronym") or "-"])
    ws2.append(["Criteria 총합", len(criteria)])
    ws2.append(["Relations 총합", sum(len(c.get("relations") or []) for c in criteria)])
    ws2.append(["검수 일시", datetime.now().strftime("%Y-%m-%d %H:%M")])

    ws2.append([])
    ws2.append(["Validator-flagged issues (자동 검출)"])
    ws2.append(["criterion_id", "scope", "issue_kind"])
    for c in criteria:
        ci = (c.get("_validation") or {}).get("issues") or []
        for issue in ci:
            ws2.append([c["criterion_id"], "criterion", issue])
        for r in c.get("relations") or []:
            ri = (r.get("_validation") or {}).get("issues") or []
            for issue in ri:
                ws2.append([c["criterion_id"], "relation:" + (r.get("relation_type") or "?"), issue])

    for col in ws2.columns:
        letter = col[0].column_letter
        ws2.column_dimensions[letter].width = 30

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  ✓ saved: {out_path}")


def fetch_neo4j_triage() -> dict:
    """Step 2 Q0.1 결과를 Neo4j에서 직접 가져옴."""
    drv = GraphDatabase.driver(os.getenv("NEO4J_URI"),
                                auth=(os.getenv("NEO4J_USER"), os.getenv("NEO4J_PASSWORD")))
    rows = []
    with drv.session(database=os.getenv("NEO4J_DATABASE", "neo4j")) as s:
        r = s.run("""
          MATCH (t:Trial)
          OPTIONAL MATCH (t)--(c:Criterion)
          WITH t, count(DISTINCT c) AS n_crit,
               sum(CASE WHEN c._passed = false THEN 1 ELSE 0 END) AS crit_fail
          OPTIONAL MATCH (t)--(:Criterion)-[r]->(:ConceptRef)
          RETURN t.nct_id AS nct, t.trial_acronym AS acronym, n_crit, crit_fail,
                 count(DISTINCT r) AS n_rel,
                 sum(CASE WHEN r._passed = false THEN 1 ELSE 0 END) AS rel_fail
          ORDER BY rel_fail DESC, crit_fail DESC
        """).data()
        for row in r:
            rows.append(row)
    drv.close()
    return {"per_trial": rows}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--trial", required=True, help="NCT id (예: NCT03425643)")
    ap.add_argument("--input", type=Path,
                    default=Path(__file__).resolve().parent / "output")
    ap.add_argument("--output", type=Path, default=None)
    args = ap.parse_args()

    json_file = args.input / f"{args.trial}_annotation.json"
    if not json_file.exists():
        print(f"ERROR: {json_file} not found", file=sys.stderr)
        sys.exit(1)
    with open(json_file, encoding="utf-8") as f:
        trial_data = json.load(f)

    print(f"  Step 2: Neo4j triage 가져오는 중...")
    neo4j_summary = fetch_neo4j_triage()
    print(f"  Step 3-5: criterion 단위 평가 ({len(trial_data['criteria'])} criteria)...")

    if args.output is None:
        out = Path(__file__).resolve().parent.parent / "results" / f"review_{args.trial}.xlsx"
    else:
        out = args.output

    build_xlsx(trial_data, neo4j_summary, out)
    print(f"\n  Trial {args.trial} 검수 1차 산출물 생성 완료.")
    print(f"  open: open '{out}'")


if __name__ == "__main__":
    main()
