"""
NSCLC Target Protocol Selection
================================
Sampling Frame(01_aact_nsclc_sampling_frame_v2.py 출력)에서
NCCN v5.2026 치료 경로 기반 stratified purposive sampling으로
타겟 프로토콜 후보를 선정하는 스크립트

입력: results/nsclc_sampling_frame.xlsx
출력: results/nsclc_protocol_candidates.xlsx

변경 이력:
  - v1 (2026-05): 초기 작성
  - v1.1 (2026-05): 품질 개선
    - SCLC, diagnostic, pan-tumor basket 시험 필터/플래그 추가
    - 셀 내 다양성 로직 (greedy diversified selection) 도입
    - NSCLC-specific 시험 우선, sponsor 다양성 반영
"""

import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 1. 설정
# ============================================================
INPUT_PATH = "../results/nsclc_sampling_frame.xlsx"
OUTPUT_PATH = "../results/nsclc_protocol_candidates.xlsx"

# 셀당 최대 후보 수 (enrollment 상위 N개)
CANDIDATES_PER_CELL = 5

# ============================================================
# 2. 품질 필터 & 플래그
# ============================================================
# v1.1: SCLC, diagnostic, pan-tumor basket 시험 필터/플래그

EXCLUDE_PATTERNS = [
    # SCLC 전용: title에 "SCLC" 단어가 있되, 같은 title에 NSCLC/non-small cell이 없는 경우
    # → regex만으로는 부정확하므로 함수에서 처리
    # 순수 diagnostic 연구
    r"\bdiagnostic\s+study\b",
    r"^\s*(?:a\s+)?(?:clinical\s+)?study\s+testing\s+dna\b",
]
EXCLUDE_REGEX = re.compile("|".join(EXCLUDE_PATTERNS), re.IGNORECASE)

# SCLC 판정은 별도 함수로 (regex 한 줄로 처리 불가)
def is_sclc_only(title):
    """SCLC 전용 시험인지 판정 (NSCLC 시험은 제외하지 않음)"""
    t = str(title).lower()
    has_sclc = bool(re.search(r"\bsclc\b", t)) or bool(re.search(r"\bsmall\s*cell\s*lung\b", t))
    has_nsclc = "nsclc" in t or "non-small" in t or "non small" in t
    return has_sclc and not has_nsclc

BASKET_PATTERNS = [
    r"\bsolid\s+tumou?rs?\b",
    r"\bpan.?tumou?r\b",
    r"\bmultiple\s+tumou?r\s+types\b",
    r"\btumou?r.?agnostic\b",
]
BASKET_REGEX = re.compile("|".join(BASKET_PATTERNS), re.IGNORECASE)


def apply_quality_filters(frame):
    """품질 필터 적용: 부적합 시험 제외 + 바스켓 시험 플래그"""

    n_before = len(frame)

    # SCLC 전용 제외
    title_text = frame["Brief_Title"].fillna("")
    sclc_mask = title_text.apply(is_sclc_only)
    n_sclc = sclc_mask.sum()

    # Diagnostic 등 패턴 제외
    pattern_mask = title_text.str.contains(EXCLUDE_REGEX)
    n_pattern = pattern_mask.sum()

    exclude_mask = sclc_mask | pattern_mask
    n_excluded = exclude_mask.sum()
    frame = frame[~exclude_mask].copy()

    # 바스켓 시험 플래그 (제외하지 않고 우선순위만 낮춤)
    title_text = frame["Brief_Title"].fillna("")
    frame["is_basket"] = title_text.str.contains(BASKET_REGEX).astype(int)

    print(f"   품질 필터: {n_excluded}건 제외 (SCLC: {n_sclc}, diagnostic 등: {n_pattern})")
    print(f"   바스켓 시험: {frame['is_basket'].sum()}건 플래그 (NSCLC-specific 우선)")

    return frame

# ============================================================
# 2. 단순화 함수
# ============================================================

# v1.2: chemoRT 맥락 키워드 — 이 키워드가 있는 시험은 Locally Adv (III)로 유지
CHEMORE_KEYWORDS = re.compile(
    r"chemoradio|concurrent.*radi|radical.*radio|definitive.*radio"
    r"|thoracic.*radi|chemo.*radiation|radiation.*chemo"
    r"|stage\s*iii[abc]?\b.*\bunresectable\b"
    r"|\bunresectable\b.*stage\s*iii[abc]?\b"
    r"|consolidat.*after.*radi|consolidat.*chemorad"
    r"|\brtog\b|\bpacific\b",
    re.IGNORECASE,
)


def simplify_stage(s, all_text=""):
    """복합 stage 태그를 primary stage로 단순화

    v1.2: chemoRT 맥락이 있는 시험은 복합 태깅이어도
    Locally Adv (III)로 유지 (Metastatic으로 흡수 방지)
    """
    s = str(s)
    if s in ("Not specified", "nan"):
        return "Not specified"
    # 단일 stage — 그대로
    if s in ("Early (I-II)", "Early (I-II) / Perioperative", "Perioperative"):
        return "Early (I-II)"
    if s == "Locally Advanced (III)":
        return "Locally Adv (III)"
    if s == "Metastatic (IV)":
        return "Metastatic (IV)"
    if s == "Advanced (III/IV)":
        return "Advanced (III/IV)"

    # 복합 stage — chemoRT 맥락 확인
    has_locally_adv = "Locally Advanced (III)" in s
    has_metastatic = "Metastatic (IV)" in s or "Advanced (III/IV)" in s
    has_chemoRT = bool(CHEMORE_KEYWORDS.search(str(all_text)))

    # chemoRT 맥락 + Locally Advanced 포함 → Locally Adv 유지
    if has_locally_adv and has_chemoRT:
        return "Locally Adv (III)"

    # 그 외: 가장 진행된 stage로
    if has_metastatic:
        if "Metastatic (IV)" in s:
            return "Metastatic (IV)"
        return "Advanced (III/IV)"
    if has_locally_adv:
        return "Locally Adv (III)"
    if "Early (I-II)" in s:
        return "Early (I-II)"
    return "Not specified"


def simplify_line(l):
    """Line of therapy 단순화"""
    l = str(l)
    if "1L + Maintenance" in l:
        return "Maint/Consol"
    if "Maintenance" in l or "Consolidation" in l:
        return "Maint/Consol"
    if "1L" in l:
        return "1L"
    if "2L+" in l:
        return "2L+"
    if "Mixed" in l:
        return "Mixed"
    return "Not specified"


def extract_primary_biomarker(b):
    """바이오마커 조합에서 primary driver 추출"""
    b = str(b)
    if b in ("Not specified", "nan"):
        return "No driver"

    markers = [m.strip() for m in b.split(",")]

    # PD-L1만 있으면 IO biomarker
    if markers == ["PD-L1"]:
        return "PD-L1 only"

    # actionable driver 중 첫 번째
    actionable_order = [
        "EGFR", "ALK", "ROS1", "KRAS", "BRAF",
        "MET", "RET", "NTRK", "HER2", "NRG1",
    ]
    for driver in actionable_order:
        if driver in markers:
            return driver

    if "PD-L1" in markers:
        return "PD-L1 only"
    return "Other"


def extract_primary_modality(m):
    """복합 modality에서 primary modality 추출"""
    m = str(m)
    if "ADC" in m and "Bispecific" not in m:
        return "ADC"
    if "Bispecific" in m:
        return "Bispecific Ab"
    if "Targeted (TKI)" in m and "Immunotherapy" not in m and "Chemotherapy" not in m:
        return "TKI"
    if "Immunotherapy" in m and "Chemotherapy" not in m and "Targeted" not in m:
        return "IO"
    if "Chemotherapy" in m and "Immunotherapy" in m:
        return "Chemo+IO"
    if "Chemotherapy" in m and "Immunotherapy" not in m:
        return "Chemo"
    if "Targeted (TKI)" in m and "Immunotherapy" in m:
        return "TKI+IO"
    if "Radiation" in m and "Surgery" in m:
        return "ChemoRT/Surgery"
    if "Surgery" in m:
        return "Surgery"
    if "Radiation" in m:
        return "Radiation"
    return "Other"


# ============================================================
# 4. Primary Axis 선정: Stage × Line (다양성 보장)
# ============================================================
def diversified_select(cell_df, n_select, max_candidates):
    """셀 내 bio/mod 다양성을 보장하는 2-pass greedy selection

    Pass 1: bio/mod가 겹치지 않는 시험을 먼저 n_select개까지 채움
    Pass 2: 나머지 슬롯을 enrollment 순으로 채워 max_candidates까지
    """
    if len(cell_df) == 0:
        return pd.DataFrame()

    # 기본 정렬: NSCLC-specific → Completed → enrollment 큰 순
    cell_df = cell_df.sort_values(
        ["is_basket", "_status_rank", "_enrollment"],
        ascending=[True, True, False],
    ).copy()

    selected_ids = set()
    selected_rows = []
    used_bios = set()
    used_mods = set()

    # Pass 1: 다양성 우선 (target 수까지)
    # 1a: 새로운 bio AND 새로운 mod를 동시에 가져오는 시험
    for _, row in cell_df.iterrows():
        if len(selected_rows) >= n_select:
            break
        bio = row["primary_bio"]
        mod = row["primary_mod"]
        is_new_bio = (bio not in used_bios) and (bio != "No driver")
        is_new_mod = mod not in used_mods
        if is_new_bio and is_new_mod:
            selected_rows.append(row)
            selected_ids.add(row["NCT_ID"])
            used_bios.add(bio)
            used_mods.add(mod)

    # 1b: 새로운 mod라도 가져오는 시험
    for _, row in cell_df.iterrows():
        if len(selected_rows) >= n_select:
            break
        if row["NCT_ID"] in selected_ids:
            continue
        mod = row["primary_mod"]
        if mod not in used_mods:
            selected_rows.append(row)
            selected_ids.add(row["NCT_ID"])
            used_bios.add(row["primary_bio"])
            used_mods.add(mod)

    # 1c: 새로운 bio라도 가져오는 시험
    for _, row in cell_df.iterrows():
        if len(selected_rows) >= n_select:
            break
        if row["NCT_ID"] in selected_ids:
            continue
        bio = row["primary_bio"]
        if (bio not in used_bios) and (bio != "No driver"):
            selected_rows.append(row)
            selected_ids.add(row["NCT_ID"])
            used_bios.add(bio)
            used_mods.add(row["primary_mod"])

    # 1d: 그래도 target 미달이면 enrollment 순으로 채움
    for _, row in cell_df.iterrows():
        if len(selected_rows) >= n_select:
            break
        if row["NCT_ID"] in selected_ids:
            continue
        selected_rows.append(row)
        selected_ids.add(row["NCT_ID"])
        used_bios.add(row["primary_bio"])
        used_mods.add(row["primary_mod"])

    # Pass 2: 나머지 슬롯 (max_candidates까지) — enrollment 순
    for _, row in cell_df.iterrows():
        if len(selected_rows) >= max_candidates:
            break
        if row["NCT_ID"] in selected_ids:
            continue
        selected_rows.append(row)
        selected_ids.add(row["NCT_ID"])

    return pd.DataFrame(selected_rows)


def select_primary(frame):
    """NCCN 치료 경로 핵심 셀에서 다양성 보장 후보 선정"""

    primary_cells = [
        ("Early (I-II)",      "1L",          "Perioperative",  3),
        ("Locally Adv (III)",  "1L",          "ChemoRT 1L",     2),
        ("Locally Adv (III)",  "Maint/Consol","Consolidation",  2),
        ("Advanced (III/IV)",  "1L",          "Advanced 1L",    3),
        ("Advanced (III/IV)",  "2L+",         "Advanced 2L+",   3),
        ("Metastatic (IV)",    "1L",          "Metastatic 1L",  4),
        ("Metastatic (IV)",    "2L+",         "Metastatic 2L+", 4),
        ("Metastatic (IV)",    "Maint/Consol","Maintenance",    2),
    ]

    results = []
    for stage, line, label, n_select in primary_cells:
        cell = frame[
            (frame["stage_simple"] == stage) &
            (frame["line_simple"] == line)
        ].copy()

        selected = diversified_select(cell, n_select, CANDIDATES_PER_CELL)
        if len(selected) > 0:
            selected = selected.copy()
            selected["selection_cell"] = label
            selected["selection_tier"] = "Primary"
            selected["selection_target_n"] = n_select
            results.append(selected)

    return pd.concat(results, ignore_index=True)


# ============================================================
# 5. Secondary Axis 보충: Biomarker × Modality
# ============================================================
def select_secondary(frame, already_selected_ids):
    """Primary에서 누락된 biomarker/modality 보충"""

    remaining = frame[~frame["NCT_ID"].isin(already_selected_ids)].copy()
    # v1.1: NSCLC-specific 우선, 바스켓은 뒤로
    remaining = remaining.sort_values(
        ["is_basket", "_status_rank", "_enrollment"],
        ascending=[True, True, False],
    )

    results = []

    # 4a. NCCN 필수 driver 보충
    nccn_drivers = [
        "EGFR", "ALK", "ROS1", "KRAS", "BRAF",
        "MET", "RET", "NTRK", "HER2", "NRG1",
    ]
    for driver in nccn_drivers:
        # 이미 선정된 시험 중 이 driver가 있는지 확인
        already_has = any(
            driver in str(b)
            for b in frame[frame["NCT_ID"].isin(already_selected_ids)]["Biomarker"]
        )
        if already_has:
            continue

        # 이 driver를 primary biomarker로 가진 시험 선정
        candidates = remaining[remaining["primary_bio"] == driver]
        if len(candidates) == 0:
            # primary가 아니어도 포함하고 있으면 선정
            candidates = remaining[
                remaining["Biomarker"].fillna("").str.contains(driver, case=False)
            ]
        if len(candidates) > 0:
            pick = candidates.head(1).copy()
            pick["selection_cell"] = f"Driver: {driver}"
            pick["selection_tier"] = "Secondary"
            pick["selection_target_n"] = 1
            results.append(pick)
            already_selected_ids = set(already_selected_ids) | set(pick["NCT_ID"])
            remaining = remaining[~remaining["NCT_ID"].isin(already_selected_ids)]

    # 4b. ADC / Bispecific Ab 전용 시험 보충
    for modality, label in [("ADC", "Modality: ADC"), ("Bispecific Ab", "Modality: Bispecific")]:
        already_has = any(
            modality in str(m)
            for m in frame[frame["NCT_ID"].isin(already_selected_ids)]["primary_mod"]
        )
        if already_has:
            continue

        candidates = remaining[remaining["primary_mod"] == modality]
        if len(candidates) > 0:
            pick = candidates.head(1).copy()
            pick["selection_cell"] = label
            pick["selection_tier"] = "Secondary"
            pick["selection_target_n"] = 1
            results.append(pick)
            already_selected_ids = set(already_selected_ids) | set(pick["NCT_ID"])
            remaining = remaining[~remaining["NCT_ID"].isin(already_selected_ids)]

    if results:
        return pd.concat(results, ignore_index=True)
    return pd.DataFrame()


# ============================================================
# 6. Coverage Check 보충: Histology, Phase
# ============================================================
def select_coverage(frame, already_selected_ids):
    """Squamous/Phase gap 보충"""

    selected_df = frame[frame["NCT_ID"].isin(already_selected_ids)]
    remaining = frame[~frame["NCT_ID"].isin(already_selected_ids)].copy()
    remaining = remaining.sort_values(
        ["is_basket", "_status_rank", "_enrollment"],
        ascending=[True, True, False],
    )

    results = []

    # 5a. Squamous 전용 시험
    has_squamous = any(
        h == "Squamous" for h in selected_df["Histology"]
    )
    if not has_squamous:
        candidates = remaining[remaining["Histology"] == "Squamous"]
        if len(candidates) > 0:
            pick = candidates.head(2).copy()
            pick["selection_cell"] = "Coverage: Squamous"
            pick["selection_tier"] = "Coverage"
            pick["selection_target_n"] = 1
            results.append(pick)
            already_selected_ids = set(already_selected_ids) | set(pick["NCT_ID"])
            remaining = remaining[~remaining["NCT_ID"].isin(already_selected_ids)]

    # 5b. Phase I 시험 (biomarker criteria가 촘촘한 패턴)
    has_phase1 = any(
        "PHASE1" == str(p) for p in selected_df["Phase"]
    )
    if not has_phase1:
        candidates = remaining[remaining["Phase"] == "PHASE1"]
        if len(candidates) > 0:
            pick = candidates.head(1).copy()
            pick["selection_cell"] = "Coverage: Phase I"
            pick["selection_tier"] = "Coverage"
            pick["selection_target_n"] = 1
            results.append(pick)

    # 5c. Phase III 시험 (prior therapy/washout criteria가 복잡한 패턴)
    has_phase3 = any(
        "PHASE3" == str(p) for p in selected_df["Phase"]
    )
    if not has_phase3:
        candidates = remaining[remaining["Phase"] == "PHASE3"]
        if len(candidates) > 0:
            pick = candidates.head(1).copy()
            pick["selection_cell"] = "Coverage: Phase III"
            pick["selection_tier"] = "Coverage"
            pick["selection_target_n"] = 1
            results.append(pick)

    if results:
        return pd.concat(results, ignore_index=True)
    return pd.DataFrame()


# ============================================================
# 6. Excel 출력 (포맷팅 포함)
# ============================================================
def export_candidates(candidates, frame, output_path):
    """후보 시험을 포맷팅된 Excel로 저장"""

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 출력 컬럼 정리
    export_cols = [
        "selection_tier", "selection_cell", "selection_target_n",
        "NCT_ID", "Brief_Title", "Phase", "Status", "Enrollment",
        "is_basket",  # v1.1
        "Modality", "Biomarker", "Line_of_Therapy", "Histology", "Stage",
        "primary_bio", "primary_mod", "stage_simple", "line_simple",
        "Lead_Sponsor", "Sponsor_Type",
    ]
    out = candidates[[c for c in export_cols if c in candidates.columns]].copy()
    out = out.sort_values(["selection_tier", "selection_cell", "NCT_ID"])

    out.to_excel(output_path, index=False, sheet_name="Candidates", engine="openpyxl")

    # 포맷팅
    wb = load_workbook(output_path)
    ws = wb["Candidates"]

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Tier별 색상
    tier_colors = {
        "Primary":   PatternFill("solid", fgColor="D6E4F0"),   # 연한 파란색
        "Secondary": PatternFill("solid", fgColor="E2EFDA"),   # 연한 초록색
        "Coverage":  PatternFill("solid", fgColor="FFF2CC"),   # 연한 노란색
    }

    data_font = Font(name="Arial", size=9)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tier = row[0].value
        fill = tier_colors.get(tier, PatternFill())
        for cell in row:
            cell.font = data_font
            cell.fill = fill

    # 열 너비
    col_widths = {
        "A": 10, "B": 22, "C": 8, "D": 14, "E": 55,
        "F": 12, "G": 18, "H": 10, "I": 30, "J": 25,
        "K": 20, "L": 14, "M": 28, "N": 12, "O": 14,
        "P": 18, "Q": 12, "R": 25, "S": 10,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 필터 추가
    ws.auto_filter.ref = ws.dimensions

    # Summary 시트 추가
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "NSCLC Protocol Selection Summary"
    ws2["A1"].font = Font(bold=True, name="Arial", size=14, color="2F5496")

    ws2["A3"] = "Selection Tier"
    ws2["B3"] = "Count"
    ws2["A3"].font = Font(bold=True, name="Arial", size=10)
    ws2["B3"].font = Font(bold=True, name="Arial", size=10)

    tier_counts = out["selection_tier"].value_counts()
    row_idx = 4
    total = 0
    for tier in ["Primary", "Secondary", "Coverage"]:
        cnt = tier_counts.get(tier, 0)
        ws2[f"A{row_idx}"] = tier
        ws2[f"B{row_idx}"] = cnt
        ws2[f"A{row_idx}"].fill = tier_colors.get(tier, PatternFill())
        ws2[f"A{row_idx}"].font = Font(name="Arial", size=10)
        ws2[f"B{row_idx}"].font = Font(name="Arial", size=10)
        total += cnt
        row_idx += 1

    ws2[f"A{row_idx}"] = "Total Candidates"
    ws2[f"B{row_idx}"] = total
    ws2[f"A{row_idx}"].font = Font(bold=True, name="Arial", size=10)
    ws2[f"B{row_idx}"].font = Font(bold=True, name="Arial", size=10)

    row_idx += 2
    ws2[f"A{row_idx}"] = "Selection Cell"
    ws2[f"B{row_idx}"] = "Candidates"
    ws2[f"C{row_idx}"] = "Target"
    ws2[f"A{row_idx}"].font = Font(bold=True, name="Arial", size=10)
    ws2[f"B{row_idx}"].font = Font(bold=True, name="Arial", size=10)
    ws2[f"C{row_idx}"].font = Font(bold=True, name="Arial", size=10)
    row_idx += 1

    cell_counts = out.groupby(["selection_tier", "selection_cell"]).agg(
        candidates=("NCT_ID", "count"),
        target_n=("selection_target_n", "first"),
    ).reset_index()

    for _, r in cell_counts.iterrows():
        ws2[f"A{row_idx}"] = f"[{r['selection_tier']}] {r['selection_cell']}"
        ws2[f"B{row_idx}"] = r["candidates"]
        ws2[f"C{row_idx}"] = r["target_n"]
        ws2[f"A{row_idx}"].fill = tier_colors.get(r["selection_tier"], PatternFill())
        ws2[f"A{row_idx}"].font = Font(name="Arial", size=9)
        ws2[f"B{row_idx}"].font = Font(name="Arial", size=9)
        ws2[f"C{row_idx}"].font = Font(name="Arial", size=9)
        row_idx += 1

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 10

    # Methods 기술 참고
    row_idx += 2
    ws2[f"A{row_idx}"] = "Methods 기술 참고"
    ws2[f"A{row_idx}"].font = Font(bold=True, name="Arial", size=11, color="2F5496")
    row_idx += 1
    methods_text = (
        "6개 층화 기준(Stage, Line of Therapy, Biomarker, Modality, "
        "Histology, Phase)에 대해 NCCN v5.2026의 치료 경로 분기 구조에 "
        "기반한 stratified purposive sampling을 수행하였다. "
        "Stage × Line of Therapy를 primary axis로 하여 각 셀에서 "
        "최소 2개 시험을 선정하고, Biomarker × Modality 다양성 및 "
        "Histology/Phase coverage를 보충하여 총 N개 시험을 선정하였다. "
        "추가 시험의 선정은 새로운 eligibility criteria 패턴이 "
        "더 이상 관찰되지 않을 때(theoretical saturation) 중단하였다."
    )
    ws2[f"A{row_idx}"] = methods_text
    ws2[f"A{row_idx}"].font = Font(name="Arial", size=9, italic=True)
    ws2[f"A{row_idx}"].alignment = Alignment(wrap_text=True)
    ws2.merge_cells(f"A{row_idx}:C{row_idx + 3}")

    wb.save(output_path)
    print(f"\n✅ 후보 시험 저장: {output_path}")
    print(f"   → {total}건 (Primary {tier_counts.get('Primary', 0)} + "
          f"Secondary {tier_counts.get('Secondary', 0)} + "
          f"Coverage {tier_counts.get('Coverage', 0)})")


# ============================================================
# 7. 메인 실행
# ============================================================
def main():
    print("=" * 60)
    print("  NSCLC Target Protocol Selection")
    print("  NCCN v5.2026 치료 경로 기반 Stratified Purposive Sampling")
    print("=" * 60)

    if not os.path.exists(INPUT_PATH):
        print(f"\n❌ 입력 파일 없음: {INPUT_PATH}")
        print(f"   먼저 01_aact_nsclc_sampling_frame_v2.py를 실행하세요.")
        return

    # 로드
    print(f"\n📂 Sampling Frame 로딩: {INPUT_PATH}")
    frame = pd.read_excel(INPUT_PATH, engine="openpyxl")
    print(f"   → {len(frame):,}건")

    # 단순화 컬럼 추가
    # v1.2: _all_text가 있으면 chemoRT 맥락 판정에 사용
    all_text_col = frame.get("_all_text", pd.Series("", index=frame.index)).fillna("")
    # _all_text가 없으면 Brief_Title로 대체
    if all_text_col.str.len().sum() == 0:
        all_text_col = frame["Brief_Title"].fillna("")
    frame["stage_simple"] = frame.apply(
        lambda r: simplify_stage(r["Stage"], all_text_col.get(r.name, "")),
        axis=1,
    )
    frame["line_simple"] = frame["Line_of_Therapy"].apply(simplify_line)
    frame["primary_bio"] = frame["Biomarker"].apply(extract_primary_biomarker)
    frame["primary_mod"] = frame["Modality"].apply(extract_primary_modality)

    # 선정 우선순위 컬럼
    status_priority = {
        "COMPLETED": 0,
        "ACTIVE_NOT_RECRUITING": 1,
        "RECRUITING": 2,
        "ENROLLING_BY_INVITATION": 3,
        "NOT_YET_RECRUITING": 4,
        "TERMINATED": 5,
        "SUSPENDED": 6,
        "WITHDRAWN": 7,
    }
    frame["_status_rank"] = frame["Status"].map(status_priority).fillna(8)
    frame["_enrollment"] = pd.to_numeric(frame["Enrollment"], errors="coerce").fillna(0)

    # v1.1: 품질 필터 적용
    print("\n🧹 품질 필터 적용 중...")
    frame = apply_quality_filters(frame)
    print(f"   → 필터 후: {len(frame):,}건")

    # Step 1: Primary Axis
    print("\n🔷 Step 1: Primary Axis (Stage × Line of Therapy)")
    primary = select_primary(frame)
    print(f"   → {len(primary)}건 후보")
    for cell in primary["selection_cell"].unique():
        n = len(primary[primary["selection_cell"] == cell])
        print(f"     - {cell}: {n}건")

    # Step 2: Secondary Axis
    print("\n🔶 Step 2: Secondary Axis (Biomarker × Modality 보충)")
    secondary = select_secondary(frame, set(primary["NCT_ID"]))
    if len(secondary) > 0:
        print(f"   → {len(secondary)}건 보충")
        for cell in secondary["selection_cell"].unique():
            n = len(secondary[secondary["selection_cell"] == cell])
            print(f"     - {cell}: {n}건")
    else:
        print("   → 보충 불필요 (Primary에서 모두 포함)")

    # Step 3: Coverage Check
    all_selected = set(primary["NCT_ID"])
    if len(secondary) > 0:
        all_selected |= set(secondary["NCT_ID"])

    print("\n🔸 Step 3: Coverage Check (Histology, Phase 보충)")
    coverage = select_coverage(frame, all_selected)
    if len(coverage) > 0:
        print(f"   → {len(coverage)}건 보충")
        for cell in coverage["selection_cell"].unique():
            n = len(coverage[coverage["selection_cell"] == cell])
            print(f"     - {cell}: {n}건")
    else:
        print("   → 보충 불필요")

    # 합치기
    parts = [primary]
    if len(secondary) > 0:
        parts.append(secondary)
    if len(coverage) > 0:
        parts.append(coverage)
    candidates = pd.concat(parts, ignore_index=True)

    # 요약
    print(f"\n{'=' * 60}")
    print(f"📋 선정 결과 요약")
    print(f"{'=' * 60}")
    print(f"  총 후보: {len(candidates)}건")
    print(f"  - Primary:   {len(primary)}건 (셀당 최대 {CANDIDATES_PER_CELL}개)")
    print(f"  - Secondary: {len(secondary)}건")
    print(f"  - Coverage:  {len(coverage)}건")
    print(f"\n  ※ 각 셀의 target_n 합계 = 최종 선정 목표 수")
    print(f"  ※ 후보 중에서 target_n개를 최종 선정하세요")
    print(f"  ※ Eligibility criteria를 확인한 뒤 saturation 도달 시 중단")

    # 내보내기
    export_candidates(candidates, frame, OUTPUT_PATH)

    # 다음 단계 안내
    print(f"\n{'=' * 60}")
    print("📋 다음 단계")
    print("=" * 60)
    print("""
1. results/nsclc_protocol_candidates.xlsx를 열어
   각 셀(selection_cell)에서 target_n개를 최종 선정하세요.

2. 선정 기준:
   - Brief_Title과 Eligibility Criteria를 읽고
     criteria 구조가 다양한 시험을 우선 선정
   - 같은 셀 안에서 sponsor가 겹치지 않도록
   - Phase 다양성 확인

3. 선정 완료 후:
   - 선정된 시험의 NCT_ID 목록을 확정
   - nsclc_sampling_frame_eligibility.xlsx에서
     해당 시험의 Eligibility Criteria를 추출
   - 두 annotator가 독립적으로 구조화 작업 진행
""")


if __name__ == "__main__":
    main()